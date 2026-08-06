"""Le classeur Excel du plan.

Un classeur ne se relit pas à l'œil dans une suite de tests : ce qui est
vérifié ici, c'est que la structure, les formats de nombre et les styles ont
bien été écrits, et surtout que les cellules restent **numériques** — un
tableau dont les nombres sont du texte n'est ni triable ni calculable, et
c'est le défaut le plus courant des exports « présentables ».
"""

from __future__ import annotations

import pytest
from openpyxl import load_workbook

from cfd_traj.core.symmetry import CalcConfig, SymmetryGroup, SymmetrySpec
from cfd_traj.data.columns import build_specs
from cfd_traj.data.study import (
    BandSpec,
    DeflectionSet,
    DoeMethod,
    DoeSpec,
    EnvelopeSpec,
    parse_study,
)
from cfd_traj.engine.bands import build_bands
from cfd_traj.engine.coverage import check_coverage
from cfd_traj.engine.doe import build_plan
from cfd_traj.engine.envelope import build_envelope
from cfd_traj.report.excel import (
    FMT_ENTIER,
    FMT_SCI,
    LIBELLE_CONFIG,
    _fr,
    write_plan_excel,
)

C4V = SymmetrySpec(group=SymmetryGroup.C4V)
FEUILLES = ("Synthèse", "Plan de calcul", "Enveloppe", "Paramètres")


@pytest.fixture
def etude():
    """Une étude minimale, suffisante pour nommer et dater le classeur."""
    return parse_study(
        {
            "etude": {"nom": "LOT_ESSAI", "source": "TRAJECTOIRES"},
            "reference": {"longueur_m": 2.5},
        }
    )


@pytest.fixture
def plan(dataset_realiste):
    """Un plan de taille raisonnable, avec trois jeux de braquages."""
    specs, _ = build_specs(dataset_realiste.columns, dataset_realiste.column_values(), {})
    bandes = build_bands(dataset_realiste.values("Mach"), BandSpec(n_bands=3, min_points=50))
    envelope = build_envelope(
        dataset_realiste, band_set=bandes, specs=specs, spec=EnvelopeSpec(), symmetry=C4V
    )
    doe = DoeSpec(
        method=DoeMethod.LHS,
        n_lhs_per_band=4,
        max_nodes=100_000,
        deflections=(
            DeflectionSet("neutre"),
            DeflectionSet("tangage", dm=15.0),
            DeflectionSet("roulis", dl=15.0),
        ),
    )
    return build_plan(envelope, doe=doe, symmetry=C4V, ds=dataset_realiste)


@pytest.fixture
def classeur(plan, etude, tmp_path):
    """Le classeur écrit puis relu."""
    cible = write_plan_excel(plan, etude, tmp_path / "PLAN.xlsx")
    return load_workbook(cible)


class TestStructure:
    def test_the_workbook_has_the_four_expected_sheets_in_order(self, classeur):
        assert classeur.sheetnames == list(FEUILLES)

    def test_the_summary_comes_first(self, classeur):
        assert classeur.active.title == "Synthèse"

    def test_the_file_is_created_with_its_parent_directory(self, plan, etude, tmp_path):
        cible = write_plan_excel(plan, etude, tmp_path / "a" / "b" / "PLAN.xlsx")

        assert cible.exists()
        assert cible.stat().st_size > 5_000

    def test_the_document_properties_describe_the_plan(self, classeur, plan):
        proprietes = classeur.properties

        assert "LOT_ESSAI" in str(proprietes.title)
        assert proprietes.creator == "cfd-traj"
        assert str(plan.n_nodes) in str(proprietes.description)


class TestPlanSheet:
    def test_one_row_per_case(self, classeur, plan):
        ws = classeur["Plan de calcul"]
        entete = _ligne_entete(ws)

        assert ws.max_row - entete == plan.n_nodes

    def test_every_variable_of_the_plan_has_its_column(self, classeur, plan):
        ws = classeur["Plan de calcul"]
        entetes = _entetes(ws)

        for nom in plan.variable_names:
            assert any(nom in e or e in ("Mach", "α_tot [°]", "φ [°]", "Re") for e in entetes)
        assert len(entetes) == 4 + len(plan.variable_names) + 8

    def test_the_headers_are_grouped(self, classeur):
        ws = classeur["Plan de calcul"]
        groupes = {
            ws.cell(row=_ligne_entete(ws) - 1, column=c).value for c in range(1, ws.max_column + 1)
        }

        assert {"Identification", "Conditions de vol", "Braquages", "Calcul"} <= groupes

    def test_numbers_are_stored_as_numbers_not_text(self, classeur):
        """Le défaut classique d'un export « présentable » : des nombres en texte."""
        ws = classeur["Plan de calcul"]
        premiere = _ligne_entete(ws) + 1
        numeriques = {"Mach min", "Mach max", "Mach", "α_tot [°]", "Re", "Coût relatif"}

        for colonne, entete in enumerate(_entetes(ws), start=1):
            if entete not in numeriques:
                continue
            for rangee in range(premiere, min(premiere + 30, ws.max_row + 1)):
                valeur = ws.cell(row=rangee, column=colonne).value
                assert isinstance(valeur, int | float), f"{entete} ligne {rangee}"

    def test_the_reynolds_column_uses_scientific_notation(self, classeur):
        ws = classeur["Plan de calcul"]
        colonne = _entetes(ws).index("Re") + 1

        assert ws.cell(row=_ligne_entete(ws) + 1, column=colonne).number_format == FMT_SCI

    def test_the_configurations_are_written_in_french(self, classeur):
        ws = classeur["Plan de calcul"]
        colonne = _entetes(ws).index("Configuration de calcul") + 1

        valeurs = {
            ws.cell(row=r, column=colonne).value
            for r in range(_ligne_entete(ws) + 1, ws.max_row + 1)
        }
        assert valeurs <= set(LIBELLE_CONFIG.values())
        assert valeurs

    def test_each_configuration_carries_its_own_shade(self, classeur):
        ws = classeur["Plan de calcul"]
        colonne = _entetes(ws).index("Configuration de calcul") + 1

        teintes = {}
        for r in range(_ligne_entete(ws) + 1, ws.max_row + 1):
            cellule = ws.cell(row=r, column=colonne)
            teintes.setdefault(cellule.value, set()).add(cellule.fill.fgColor.rgb)

        assert all(len(v) == 1 for v in teintes.values())
        assert len({next(iter(v)) for v in teintes.values()}) >= 2

    def test_the_table_is_filterable_and_the_header_stays_visible(self, classeur):
        ws = classeur["Plan de calcul"]

        assert ws.auto_filter.ref is not None
        assert ws.auto_filter.ref.startswith(f"A{_ligne_entete(ws)}")
        assert ws.freeze_panes is not None

    def test_it_is_ready_to_print(self, classeur):
        ws = classeur["Plan de calcul"]

        assert ws.page_setup.orientation == "landscape"
        assert ws.sheet_properties.pageSetUpPr.fitToPage
        assert ws.print_title_rows  # l'en-tête se répète sur chaque page

    def test_every_column_has_an_explicit_width(self, classeur):
        from openpyxl.utils import get_column_letter

        ws = classeur["Plan de calcul"]

        for c in range(1, ws.max_column + 1):
            largeur = ws.column_dimensions[get_column_letter(c)].width
            assert largeur and largeur > 5.0


class TestSummarySheet:
    def test_it_answers_the_headline_questions(self, classeur):
        intitules = _intitules(classeur["Synthèse"])

        for attendu in (
            "Nombre de cas de calcul",
            "Coût total (équivalents configuration complète)",
            "Économie apportée par les symétries",
            "Groupe de symétrie",
        ):
            assert attendu in intitules

    def test_the_figures_match_the_plan(self, classeur, plan):
        valeurs = _valeurs(classeur["Synthèse"])

        assert valeurs["Nombre de cas de calcul"] == plan.n_nodes
        assert valeurs["Coût total (équivalents configuration complète)"] == pytest.approx(
            plan.total_cost
        )
        assert valeurs["Économie apportée par les symétries"] == pytest.approx(plan.saving)

    def test_a_seed_is_not_written_as_a_quantity(self, classeur):
        """Une graine est un identifiant : la grouper par milliers n'a pas de sens."""
        ws = classeur["Synthèse"]

        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "Graine de reproductibilité":
                assert ws.cell(row=r, column=2).number_format != FMT_ENTIER
                return
        pytest.fail("la graine n'apparaît pas dans la synthèse")

    def test_the_coverage_block_appears_only_when_it_was_computed(
        self, plan, etude, dataset_realiste, tmp_path
    ):
        sans = load_workbook(write_plan_excel(plan, etude, tmp_path / "sans.xlsx"))
        assert "Contrôle de couverture" not in _intitules(sans["Synthèse"])

        resultat = check_coverage(dataset_realiste, envelope=plan.envelope)
        avec = load_workbook(
            write_plan_excel(plan, etude, tmp_path / "avec.xlsx", coverage=resultat)
        )
        valeurs = _valeurs(avec["Synthèse"])
        assert valeurs["Taux de couverture des trajectoires"] == pytest.approx(resultat.rate)

    def test_the_coverage_rate_is_shown_precisely_enough_to_be_honest(
        self, plan, etude, dataset_realiste, tmp_path
    ):
        """À une décimale, 99,96 % s'afficherait « 100,0 % » et mentirait."""
        resultat = check_coverage(dataset_realiste, envelope=plan.envelope)
        wb = load_workbook(write_plan_excel(plan, etude, tmp_path / "c.xlsx", coverage=resultat))
        ws = wb["Synthèse"]

        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "Taux de couverture des trajectoires":
                assert ws.cell(row=r, column=2).number_format.startswith("0.00")
                return
        pytest.fail("le taux de couverture n'apparaît pas")


class TestEnvelopeSheet:
    def test_one_row_per_band_and_variable(self, classeur, plan):
        ws = classeur["Enveloppe"]
        attendu = sum(len(b.variables) for b in plan.envelope.bands)

        assert ws.max_row - _ligne_entete(ws) == attendu

    def test_the_bounds_stay_numeric(self, classeur):
        ws = classeur["Enveloppe"]
        colonne = _entetes(ws).index("Borne basse") + 1

        for r in range(_ligne_entete(ws) + 1, min(_ligne_entete(ws) + 20, ws.max_row + 1)):
            assert isinstance(ws.cell(row=r, column=colonne).value, int | float)

    def test_the_roles_are_written_in_french(self, classeur):
        ws = classeur["Enveloppe"]
        colonne = _entetes(ws).index("Rôle") + 1

        valeurs = {
            ws.cell(row=r, column=colonne).value
            for r in range(_ligne_entete(ws) + 1, ws.max_row + 1)
        }
        assert "dimension principale" in valeurs


class TestParametersSheet:
    def test_every_active_variable_is_listed(self, classeur, plan):
        ws = classeur["Paramètres"]
        listees = {ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)}

        for spec in plan.envelope.specs:
            if spec.is_active:
                assert spec.display in listees

    def test_the_configuration_legend_carries_every_cost(self, classeur):
        ws = classeur["Paramètres"]
        libelles = {ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)}

        for config in CalcConfig:
            assert LIBELLE_CONFIG[config] in libelles


class TestFrenchNumbers:
    @pytest.mark.parametrize(
        ("valeur", "attendu"),
        [(22.5, "22,5"), (45.0, "45"), (0.0, "0"), (0.1, "0,1"), (1200.0, "1200")],
    )
    def test_inline_numbers_use_a_decimal_comma(self, valeur, attendu):
        assert _fr(valeur) == attendu

    def test_no_sheet_leaks_a_decimal_point_into_its_prose(self, classeur):
        """Les nombres composés dans du texte ne sont pas localisés par Excel."""
        import re

        suspect = re.compile(r"\d+\.\d")
        for nom in FEUILLES:
            ws = classeur[nom]
            for rangee in ws.iter_rows(max_row=min(ws.max_row, 60)):
                for cellule in rangee:
                    if isinstance(cellule.value, str):
                        assert not suspect.search(cellule.value), f"{nom}: {cellule.value!r}"


class TestGenericity:
    @pytest.mark.parametrize("n_extra", [0, 1, 5])
    def test_any_number_of_generic_columns_is_exported(self, make_lot, etude, tmp_path, n_extra):
        from cfd_traj.core.adim import Reference
        from cfd_traj.data.dataset import load_dataset
        from cfd_traj.data.derive import add_derived_columns

        extra = tuple(f"COL{i}" for i in range(n_extra))
        ds = add_derived_columns(
            load_dataset(make_lot(n_shots=3, extra=extra)),
            reference=Reference(length_m=2.5),
            symmetry=C4V,
        )
        specs, _ = build_specs(ds.columns, ds.column_values(), {})
        bandes = build_bands(ds.values("Mach"), BandSpec(n_bands=2, min_points=10))
        envelope = build_envelope(
            ds, band_set=bandes, specs=specs, spec=EnvelopeSpec(), symmetry=C4V
        )
        plan = build_plan(
            envelope,
            doe=DoeSpec(method=DoeMethod.LHS, n_lhs_per_band=3, max_nodes=100_000),
            symmetry=C4V,
            ds=ds,
        )

        wb = load_workbook(write_plan_excel(plan, etude, tmp_path / f"p{n_extra}.xlsx"))

        entetes = _entetes(wb["Plan de calcul"])
        for nom in extra:
            assert nom in entetes


# --- petites aides ----------------------------------------------------------


def _ligne_entete(ws) -> int:
    """La ligne des en-têtes de colonne : la première qui porte « Bande »."""
    for r in range(1, 12):
        for c in range(1, min(ws.max_column, 20) + 1):
            if ws.cell(row=r, column=c).value in ("Bande", "N° cas"):
                return r
    raise AssertionError("en-tête introuvable")


def _entetes(ws) -> list[str]:
    """Les en-têtes de colonne, dans l'ordre."""
    ligne = _ligne_entete(ws)
    return [
        str(ws.cell(row=ligne, column=c).value)
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=ligne, column=c).value is not None
    ]


def _intitules(ws) -> set[str]:
    """Les intitulés de la colonne de gauche d'une feuille clé/valeur."""
    return {
        str(ws.cell(row=r, column=1).value)
        for r in range(1, ws.max_row + 1)
        if ws.cell(row=r, column=1).value is not None
    }


def _valeurs(ws) -> dict[str, object]:
    """Le couple intitulé → valeur d'une feuille clé/valeur."""
    out: dict[str, object] = {}
    for r in range(1, ws.max_row + 1):
        cle = ws.cell(row=r, column=1).value
        valeur = ws.cell(row=r, column=2).value
        if cle is not None and valeur is not None:
            out[str(cle)] = valeur
    return out
