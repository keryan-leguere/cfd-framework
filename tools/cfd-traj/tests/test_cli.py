"""L'interface en ligne de commande, appelée en processus.

Aucune commande ne doit jamais laisser filer une trace d'appels : le public est
un ingénieur CFD, et un « Traceback » dans la sortie d'erreur est un défaut.
"""

from __future__ import annotations

import pandas as pd
import pytest

from cfd_traj.cli.main import EXIT_ACTION_REQUIRED, EXIT_OK, main


def _generer(tmp_path, **kw):
    """Un petit lot, prêt pour les commandes suivantes."""
    directory = tmp_path / "TRAJ"
    argv = [
        "generer",
        "--sortie",
        str(directory),
        "--n-tirs",
        str(kw.pop("n_tirs", 6)),
        "--graine",
        "1",
    ]
    for key, value in kw.items():
        argv += [f"--{key.replace('_', '-')}", str(value)]
    assert main(argv) == EXIT_OK
    return directory


class TestGenerer:
    def test_it_writes_the_requested_number_of_shots(self, tmp_path, capsys):
        directory = _generer(tmp_path, n_tirs=5)

        assert len(list(directory.glob("tir_*.csv"))) == 5
        assert (directory / "ETUDE.yaml").exists()
        assert "5 tirs" in capsys.readouterr().out

    @pytest.mark.parametrize("n", [0, 1, 4])
    def test_any_number_of_generic_columns_works(self, tmp_path, n):
        directory = tmp_path / "TRAJ"

        code = main(
            ["generer", "--sortie", str(directory), "--n-tirs", "2", "--n-parametres", str(n)]
        )

        assert code == EXIT_OK
        header = next(directory.glob("tir_*.csv")).read_text().splitlines()[0]
        assert len(header.split(",")) == 8 + n

    def test_columns_can_be_named_explicitly(self, tmp_path):
        directory = tmp_path / "TRAJ"

        main(
            [
                "generer",
                "--sortie",
                str(directory),
                "--n-tirs",
                "2",
                "--parametres",
                "PRESSION:rampe,TEMPERATURE:plateau_bruite",
            ]
        )

        header = next(directory.glob("tir_*.csv")).read_text().splitlines()[0]
        assert header.endswith("PRESSION,TEMPERATURE")

    def test_an_unknown_archetype_lists_the_valid_ones(self, tmp_path, capsys):
        with pytest.raises(SystemExit) as excinfo:
            main(["generer", "--sortie", str(tmp_path / "T"), "--parametres", "A:inexistant"])

        assert excinfo.value.code == 1
        err = capsys.readouterr().err
        assert "rampe" in err
        assert "Traceback" not in err

    def test_it_refuses_to_write_over_an_existing_lot(self, tmp_path, capsys):
        directory = _generer(tmp_path, n_tirs=2)

        with pytest.raises(SystemExit):
            main(["generer", "--sortie", str(directory), "--n-tirs", "2"])

        assert "Traceback" not in capsys.readouterr().err

    def test_the_companion_study_can_be_suppressed(self, tmp_path):
        directory = tmp_path / "TRAJ"

        main(["generer", "--sortie", str(directory), "--n-tirs", "2", "--sans-etude"])

        assert not (directory / "ETUDE.yaml").exists()

    def test_the_companion_study_survives_the_tool_writing_next_to_the_shots(self, tmp_path):
        # A directory source would later swallow the plan as if it were a shot.
        directory = _generer(tmp_path, n_tirs=3)
        (directory / "PLAN.csv").write_text("node_id,bande\nB00-N0000,0\n")

        assert main(["analyser", str(directory / "ETUDE.yaml")]) == EXIT_OK


class TestInspecter:
    def test_it_describes_the_lot(self, tmp_path, capsys):
        directory = _generer(tmp_path)

        assert main(["inspecter", str(directory)]) == EXIT_OK
        out = capsys.readouterr().out
        assert "dimension intrinsèque" in out
        assert "tirs" in out

    def test_the_analysis_can_be_skipped(self, tmp_path, capsys):
        directory = _generer(tmp_path)

        assert main(["inspecter", str(directory), "--sans-acp"]) == EXIT_OK
        assert "dimension intrinsèque" not in capsys.readouterr().out

    def test_it_accepts_a_study_file_as_source(self, tmp_path):
        directory = _generer(tmp_path)

        assert main(["inspecter", str(directory / "ETUDE.yaml")]) == EXIT_OK

    def test_it_can_export_the_statistics(self, tmp_path):
        directory = _generer(tmp_path)
        target = tmp_path / "STATS.csv"

        main(["inspecter", str(directory), "--csv", str(target)])

        assert "variable" in pd.read_csv(target).columns

    def test_it_can_write_a_figure(self, tmp_path):
        directory = _generer(tmp_path)
        target = tmp_path / "acp.png"

        main(["inspecter", str(directory), "--figure", str(target)])

        assert target.stat().st_size > 10_000

    def test_it_can_offer_the_parameters_block(self, tmp_path, capsys):
        directory = _generer(tmp_path)

        main(["inspecter", str(directory), "--proposer"])

        assert "parametres:" in capsys.readouterr().out

    def test_a_missing_source_fails_cleanly(self, tmp_path, capsys):
        with pytest.raises(SystemExit) as excinfo:
            main(["inspecter", str(tmp_path / "nulle-part")])

        assert excinfo.value.code == 1
        err = capsys.readouterr().err
        assert "introuvable" in err
        assert "Traceback" not in err

    def test_a_lot_missing_a_required_column_names_it(self, tmp_path, capsys):
        directory = tmp_path / "BAD"
        directory.mkdir()
        (directory / "tir.csv").write_text("time,Mach,Altitude,alpha\n0,0.5,100,1\n")

        with pytest.raises(SystemExit):
            main(["inspecter", str(directory)])

        err = capsys.readouterr().err
        assert "beta" in err
        assert "Traceback" not in err


class TestAnalyser:
    def test_it_builds_the_envelope(self, tmp_path, capsys):
        directory = _generer(tmp_path)

        assert main(["analyser", str(directory / "ETUDE.yaml")]) == EXIT_OK
        out = capsys.readouterr().out
        assert "Enveloppe conditionnelle" in out
        assert "Symétrie" in out

    def test_it_can_export_the_table_and_the_figure(self, tmp_path):
        directory = _generer(tmp_path)
        csv = tmp_path / "ENV.csv"
        figure = tmp_path / "env.png"

        main(
            ["analyser", str(directory / "ETUDE.yaml"), "--csv", str(csv), "--figure", str(figure)]
        )

        assert {"bande", "variable", "borne_basse"} <= set(pd.read_csv(csv).columns)
        assert figure.stat().st_size > 10_000

    def test_the_verbose_report_is_longer(self, tmp_path, capsys):
        directory = _generer(tmp_path)

        main(["analyser", str(directory / "ETUDE.yaml")])
        short = len(capsys.readouterr().out)
        main(["analyser", str(directory / "ETUDE.yaml"), "-v"])
        long = len(capsys.readouterr().out)

        assert long > short


class TestDoe:
    def test_the_latin_hypercube_produces_a_plan(self, tmp_path, capsys):
        directory = _generer(tmp_path)
        target = tmp_path / "PLAN.csv"

        code = main(
            ["doe", str(directory / "ETUDE.yaml"), "--methode", "lhs", "--sortie", str(target)]
        )

        assert code == EXIT_OK
        assert "cas de calcul" in capsys.readouterr().out
        assert len(pd.read_csv(target)) > 0

    def test_an_oversized_plan_asks_for_an_action(self, tmp_path, capsys):
        directory = _generer(tmp_path)

        code = main(["doe", str(directory / "ETUDE.yaml"), "--noeuds-max", "5"])

        assert code == EXIT_ACTION_REQUIRED
        err = capsys.readouterr().err
        assert "--methode lhs" in err
        assert "Traceback" not in err

    def test_it_can_chain_the_coverage_check(self, tmp_path, capsys):
        directory = _generer(tmp_path)

        main(["doe", str(directory / "ETUDE.yaml"), "--methode", "lhs", "--couverture"])

        out = capsys.readouterr().out
        assert "cas de calcul" in out
        assert "interpolation stricte" in out

    def test_it_can_export_yaml_and_a_figure(self, tmp_path):
        directory = _generer(tmp_path)
        yaml_target = tmp_path / "PLAN.yaml"
        figure = tmp_path / "plan.png"

        main(
            [
                "doe",
                str(directory / "ETUDE.yaml"),
                "--methode",
                "lhs",
                "--yaml",
                str(yaml_target),
                "--figure",
                str(figure),
            ]
        )

        assert "bandes" in yaml_target.read_text()
        assert figure.stat().st_size > 10_000

    def test_it_can_write_the_review_workbook(self, tmp_path):
        from openpyxl import load_workbook

        directory = _generer(tmp_path)
        cible = tmp_path / "PLAN.xlsx"

        code = main(
            [
                "doe",
                str(directory / "ETUDE.yaml"),
                "--methode",
                "lhs",
                "--excel",
                str(cible),
            ]
        )

        assert code == EXIT_OK
        wb = load_workbook(cible)
        assert wb.sheetnames == ["Synthèse", "Plan de calcul", "Enveloppe", "Paramètres"]

    def test_the_workbook_lands_next_to_the_csv_when_no_path_is_given(self, tmp_path):
        directory = _generer(tmp_path)
        csv = tmp_path / "SORTIE" / "PLAN.csv"

        assert (
            main(
                [
                    "doe",
                    str(directory / "ETUDE.yaml"),
                    "--methode",
                    "lhs",
                    "--sortie",
                    str(csv),
                    "--excel",
                ]
            )
            == EXIT_OK
        )

        assert csv.with_suffix(".xlsx").exists()

    def test_the_workbook_carries_the_coverage_without_asking_for_it(self, tmp_path):
        """Le classeur doit se suffire à lui-même, --couverture ou non."""
        from openpyxl import load_workbook

        directory = _generer(tmp_path)
        cible = tmp_path / "PLAN.xlsx"

        main(["doe", str(directory / "ETUDE.yaml"), "--methode", "lhs", "--excel", str(cible)])

        ws = load_workbook(cible)["Synthèse"]
        intitules = {ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)}
        assert "Taux de couverture des trajectoires" in intitules

    def test_no_workbook_is_written_unless_asked(self, tmp_path):
        directory = _generer(tmp_path)
        csv = tmp_path / "SORTIE" / "PLAN.csv"

        main(["doe", str(directory / "ETUDE.yaml"), "--methode", "lhs", "--sortie", str(csv)])

        assert not csv.with_suffix(".xlsx").exists()

    def test_the_seed_makes_it_reproducible(self, tmp_path):
        directory = _generer(tmp_path)
        first = tmp_path / "A.csv"
        second = tmp_path / "B.csv"

        for target in (first, second):
            main(
                [
                    "doe",
                    str(directory / "ETUDE.yaml"),
                    "--methode",
                    "lhs",
                    "--graine",
                    "77",
                    "--sortie",
                    str(target),
                ]
            )

        assert first.read_bytes() == second.read_bytes()

    def test_corners_can_be_switched_off(self, tmp_path):
        directory = _generer(tmp_path)
        target = tmp_path / "PLAN.csv"

        main(
            [
                "doe",
                str(directory / "ETUDE.yaml"),
                "--methode",
                "lhs",
                "--sans-coins",
                "--sortie",
                str(target),
            ]
        )

        assert "coin" not in set(pd.read_csv(target)["origine"])

    def test_an_invalid_method_is_rejected_by_the_parser(self, tmp_path):
        directory = _generer(tmp_path)

        with pytest.raises(SystemExit):
            main(["doe", str(directory / "ETUDE.yaml"), "--methode", "krigeage"])


class TestCouverture:
    def test_a_full_range_envelope_covers_everything(self, tmp_path, capsys):
        directory = _generer(tmp_path)
        study = directory / "ETUDE.yaml"
        study.write_text(
            study.read_text()
            .replace("quantile_bas: 0.001", "quantile_bas: 0.0")
            .replace("quantile_haut: 0.999", "quantile_haut: 1.0")
        )

        code = main(["couverture", str(study)])

        assert code == EXIT_OK
        assert "100,00 %" in capsys.readouterr().out

    def test_tight_quantiles_ask_for_an_action(self, tmp_path, capsys):
        directory = _generer(tmp_path)
        study = directory / "ETUDE.yaml"
        study.write_text(
            study.read_text()
            .replace("quantile_bas: 0.001", "quantile_bas: 0.2")
            .replace("quantile_haut: 0.999", "quantile_haut: 0.8")
            .replace("marge: 0.05", "marge: 0.0")
        )

        code = main(["couverture", str(study), "--pires", "3"])

        assert code == EXIT_ACTION_REQUIRED
        assert "les plus éloignés" in capsys.readouterr().out

    def test_it_can_export_the_offenders_and_a_figure(self, tmp_path):
        directory = _generer(tmp_path)
        csv = tmp_path / "HORS.csv"
        figure = tmp_path / "cov.png"

        main(
            [
                "couverture",
                str(directory / "ETUDE.yaml"),
                "--csv",
                str(csv),
                "--figure",
                str(figure),
            ]
        )

        assert "tir" in pd.read_csv(csv).columns
        assert figure.stat().st_size > 10_000


class TestExample:
    def test_it_copies_a_runnable_study(self, tmp_path, capsys):
        destination = tmp_path / "EX"

        assert main(["example", "--output", str(destination)]) == EXIT_OK
        assert (destination / "ETUDE.yaml").exists()
        assert "Exemple copié" in capsys.readouterr().out

    def test_the_copy_runs(self, tmp_path):
        destination = tmp_path / "EX"
        main(["example", "--output", str(destination)])

        assert main(["analyser", str(destination / "ETUDE.yaml")]) == EXIT_OK

    def test_no_generated_figure_is_copied(self, tmp_path):
        destination = tmp_path / "EX"

        main(["example", "--output", str(destination)])

        assert not list(destination.rglob("*.png"))

    def test_it_refuses_a_non_empty_destination(self, tmp_path, capsys):
        destination = tmp_path / "EX"
        destination.mkdir()
        (destination / "garde.txt").write_text("ne pas écraser")

        with pytest.raises(SystemExit):
            main(["example", "--output", str(destination)])

        assert (destination / "garde.txt").read_text() == "ne pas écraser"
        assert "Traceback" not in capsys.readouterr().err


class TestErrorHandling:
    @pytest.mark.parametrize("command", ["analyser", "doe", "couverture"])
    def test_a_missing_study_fails_cleanly(self, tmp_path, capsys, command):
        with pytest.raises(SystemExit) as excinfo:
            main([command, str(tmp_path / "absent.yaml")])

        assert excinfo.value.code == 1
        err = capsys.readouterr().err
        assert "introuvable" in err
        assert "Traceback" not in err

    @pytest.mark.parametrize("command", ["analyser", "doe", "couverture"])
    def test_a_malformed_study_fails_cleanly(self, tmp_path, capsys, command):
        study = tmp_path / "ETUDE.yaml"
        study.write_text("etude: [\n  nom: x\n")

        with pytest.raises(SystemExit):
            main([command, str(study)])

        assert "Traceback" not in capsys.readouterr().err

    def test_an_unknown_symmetry_group_lists_the_five(self, tmp_path, capsys):
        directory = _generer(tmp_path)
        study = directory / "ETUDE.yaml"
        study.write_text(study.read_text().replace("groupe: C4v", "groupe: D4h"))

        with pytest.raises(SystemExit):
            main(["analyser", str(study)])

        err = capsys.readouterr().err
        assert "C4v" in err and "Cinfv" in err
        assert "Traceback" not in err

    def test_declaring_a_column_that_is_not_there_names_the_ones_that_are(self, tmp_path, capsys):
        directory = _generer(tmp_path)
        study = directory / "ETUDE.yaml"
        study.write_text(study.read_text() + "\nparametres:\n  ABSENTE: { role: principal }\n")

        with pytest.raises(SystemExit):
            main(["analyser", str(study)])

        err = capsys.readouterr().err
        assert "ABSENTE" in err
        assert "Traceback" not in err

    def test_no_subcommand_is_refused(self):
        with pytest.raises(SystemExit):
            main([])
