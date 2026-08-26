"""Classeur Excel du plan de calcul, en français.

C'est le livrable que l'on pose sur la table en revue de définition, et celui
qu'on envoie à qui lancera les calculs. Il doit donc être lisible sans mode
d'emploi : quatre feuilles, dans l'ordre où on se pose les questions.

``Synthèse``
    Combien de calculs, pour quel coût, sous quelles hypothèses. Ce que le chef
    de projet lit, et rien d'autre.
``Plan de calcul``
    Une ligne par cas, en-têtes groupés, filtres, volets figés, mise en page
    d'impression prête. C'est le tableau de travail.
``Enveloppe``
    Les bornes conditionnelles bande par bande — la justification du plan.
``Paramètres``
    Le rôle de chaque colonne et la légende des configurations de calcul.

Les cellules restent **numériques** : les nombres sont écrits tels quels et
l'affichage est confié aux formats de nombre d'Excel. Le tableau reste donc
triable, filtrable et calculable, et il s'affiche avec les séparateurs français
sur un poste français sans qu'on ait rien codé en dur.
"""

from __future__ import annotations

from collections.abc import Callable, Mapping, Sequence
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet

from cfd_traj._compat import zip_strict
from cfd_traj.core.symmetry import CalcConfig, SymmetryGroup, azimuth_levels
from cfd_traj.data.columns import ColumnSpec, Role
from cfd_traj.data.study import Study
from cfd_traj.engine.coverage import CoverageResult
from cfd_traj.engine.doe import DoeNode, DoePlan

# --- charte graphique ------------------------------------------------------

#: Bleu ardoise sobre : lisible en noir et blanc comme en couleur.
BLEU_FONCE = "1F3B53"
BLEU_MOYEN = "2E5C7E"
BLEU_CLAIR = "DCE6EE"
GRIS_BANDE = "F4F6F8"
GRIS_TRAIT = "BFC9D2"
BLANC = "FFFFFF"
NOIR = "1A1A1A"

#: Un vert, un ambre, un rouge assez désaturés pour rester imprimables.
VERT = "E3F0E4"
AMBRE = "FCF2DC"
ROUGE = "FBE4E4"

POLICE = "Calibri"

TITRE = Font(name=POLICE, size=16, bold=True, color=BLEU_FONCE)
SOUS_TITRE = Font(name=POLICE, size=10, color="5A6A78")
ENTETE_GROUPE = Font(name=POLICE, size=10, bold=True, color=BLANC)
ENTETE = Font(name=POLICE, size=10, bold=True, color=BLANC)
CORPS = Font(name=POLICE, size=10, color=NOIR)
CORPS_GRAS = Font(name=POLICE, size=10, bold=True, color=NOIR)
LEGENDE = Font(name=POLICE, size=9, italic=True, color="5A6A78")

REMPLI_GROUPE = PatternFill("solid", fgColor=BLEU_FONCE)
REMPLI_ENTETE = PatternFill("solid", fgColor=BLEU_MOYEN)
REMPLI_BANDE = PatternFill("solid", fgColor=GRIS_BANDE)
REMPLI_ACCENT = PatternFill("solid", fgColor=BLEU_CLAIR)

_TRAIT = Side(style="thin", color=GRIS_TRAIT)
BORDURE = Border(left=_TRAIT, right=_TRAIT, top=_TRAIT, bottom=_TRAIT)

CENTRE = Alignment(horizontal="center", vertical="center", wrap_text=True)
GAUCHE = Alignment(horizontal="left", vertical="center")
DROITE = Alignment(horizontal="right", vertical="center")

# --- formats de nombre ------------------------------------------------------

#: Codes de format Excel, écrits dans la grammaire canonique : la virgule y
#: désigne le groupement des milliers et le point la décimale, quelle que soit
#: la langue. C'est Excel qui les rend ensuite avec les séparateurs de la
#: locale du poste — espace et virgule sur un poste français. Écrire une espace
#: littérale ici ne grouperait rien : ce n'est pas un séparateur valide.
FMT_ENTIER = "#,##0"
FMT_COUT = "#,##0.0"
FMT_IDENT = "0"
FMT_2 = "0.00"
FMT_3 = "0.000"
FMT_4 = "0.0000"
FMT_SCI = "0.00E+00"
# L'espace insécable avant le signe est la règle typographique française ; les
# guillemets en font une chaîne littérale dans le code de format.
FMT_PCT = '0.0" %"'
# La couverture se lit à deux décimales : à une seule, 99,96 % s'affiche
# « 100,0 % » et contredit l'état affiché juste en dessous.
FMT_PCT2 = '0.00" %"'

#: Au-delà, une colonne passe en notation scientifique : un Reynolds à huit
#: chiffres écrit en toutes lettres déborde et ne se compare pas d'un coup d'œil.
SEUIL_SCIENTIFIQUE = 1e5


@dataclass(frozen=True)
class Colonne:
    """Une colonne du tableau : d'où elle vient, comment elle s'affiche."""

    cle: str
    entete: str
    groupe: str
    format: str = FMT_3
    largeur: float = 12.0
    alignement: Alignment = DROITE


#: Libellés français des configurations de calcul, du moins cher au plus cher.
LIBELLE_CONFIG: dict[CalcConfig, str] = {
    CalcConfig.AXI_2D: "axisymétrique 2D",
    CalcConfig.SECTEUR_45: "secteur 45°",
    CalcConfig.QUART_90: "quart 90° cyclique",
    CalcConfig.DEMI: "demi-configuration",
    CalcConfig.COMPLETE: "configuration complète",
}

#: Teinte de fond par configuration : le coût se lit sans lire le chiffre.
COULEUR_CONFIG: dict[CalcConfig, str] = {
    CalcConfig.AXI_2D: VERT,
    CalcConfig.SECTEUR_45: VERT,
    CalcConfig.QUART_90: AMBRE,
    CalcConfig.DEMI: AMBRE,
    CalcConfig.COMPLETE: ROUGE,
}

LIBELLE_ROLE: dict[Role, str] = {
    Role.PRINCIPAL: "dimension principale",
    Role.CONDITIONNEL: "conditionnée au Mach",
    Role.DISCRET: "facteur discret",
    Role.MECANIQUE: "plage mécanique",
    Role.IGNORE: "exclue de l'analyse",
}

LIBELLE_GROUPE: dict[SymmetryGroup, str] = {
    SymmetryGroup.CINFV: "corps de révolution",
    SymmetryGroup.C4V: "cruciforme, quatre plans de miroir",
    SymmetryGroup.C4: "axe d'ordre 4 seul, aucun miroir",
    SymmetryGroup.CS: "un unique plan de miroir",
    SymmetryGroup.C1: "aucune symétrie",
}

#: En-têtes des variables dérivées. Les colonnes génériques prennent le nom et
#: l'unité que l'étude leur donne — rien n'est deviné à partir du nom.
ENTETE_DERIVEE: dict[str, str] = {
    "Mach": "Mach",
    "alpha_tot": "α_tot [°]",
    "phi_fold": "φ [°]",
    "Re_ref": "Re",
}


def _fr(valeur: float, decimales: int = 1) -> str:
    """Nombre à la française, pour les valeurs écrites *dans* une chaîne.

    Les cellules numériques sont localisées par Excel lui-même ; celles qui
    portent du texte ne le sont pas, et doivent donc être composées ici.
    """
    return f"{valeur:.{decimales}f}".rstrip("0").rstrip(".").replace(".", ",") or "0"


#: Colonnes en degrés : deux décimales suffisent, et un dixième de degré de
#: plus n'apporte rien à un cas de calcul.
COLONNES_ANGULAIRES: frozenset[str] = frozenset({"alpha_tot", "phi_fold"})


def _format_variable(spec: ColumnSpec | None, nom: str, echelle: float) -> str:
    """Format de nombre adapté à l'amplitude et au rôle de la variable."""
    if nom == "Re_ref" or echelle >= SEUIL_SCIENTIFIQUE:
        return FMT_SCI
    if nom == "Mach":
        return FMT_3
    if nom in COLONNES_ANGULAIRES or (spec is not None and spec.unit == "deg"):
        return FMT_2
    return FMT_3


def _entete_variable(spec: ColumnSpec | None, nom: str) -> str:
    """En-tête d'une colonne variable : nom, unité, libellé déclarés."""
    if nom in ENTETE_DERIVEE:
        return ENTETE_DERIVEE[nom]
    if spec is None:
        return nom
    if spec.unit:
        return f"{nom} [{spec.unit}]"
    return nom


def _colonnes(plan: DoePlan) -> list[Colonne]:
    """Le modèle de colonnes du tableau, dérivé du plan et non codé en dur."""
    specs = {s.name: s for s in plan.envelope.specs}
    colonnes = [
        Colonne("node_id", "N° cas", "Identification", "@", 13.0, GAUCHE),
        Colonne("bande", "Bande", "Identification", FMT_ENTIER, 7.0, CENTRE),
        Colonne("mach_bas", "Mach min", "Identification", FMT_2, 10.0),
        Colonne("mach_haut", "Mach max", "Identification", FMT_2, 10.0),
    ]

    for nom in plan.variable_names:
        spec = specs.get(nom)
        echelle = _amplitude(plan, nom)
        colonnes.append(
            Colonne(
                nom,
                _entete_variable(spec, nom),
                "Conditions de vol",
                _format_variable(spec, nom, echelle),
                13.0,
            )
        )

    colonnes.extend(
        [
            Colonne("braquage", "Jeu", "Braquages", "@", 12.0, GAUCHE),
            Colonne("dl", "δl [°]", "Braquages", FMT_2, 9.0),
            Colonne("dm", "δm [°]", "Braquages", FMT_2, 9.0),
            Colonne("dn", "δn [°]", "Braquages", FMT_2, 9.0),
            Colonne("configuration", "Configuration de calcul", "Calcul", "@", 22.0, GAUCHE),
            Colonne("cout_relatif", "Coût relatif", "Calcul", FMT_3, 11.0),
            Colonne("composantes_nulles", "Composantes nulles", "Calcul", "@", 18.0, CENTRE),
            Colonne("origine", "Origine", "Calcul", "@", 10.0, CENTRE),
        ]
    )
    return colonnes


def _amplitude(plan: DoePlan, nom: str) -> float:
    """Ordre de grandeur d'une variable sur tout le plan."""
    valeurs = [abs(n.values[nom]) for n in plan.nodes if nom in n.values]
    finis = [v for v in valeurs if v == v]
    return max(finis) if finis else 0.0


#: Colonnes calculées, par clé française. Toute clé absente de cette table est
#: une variable du plan, lue telle quelle dans ``node.values``.
_COLONNES_CALCULEES: Mapping[str, Callable[[DoeNode], Any]] = {
    "node_id": lambda node: node.node_id,
    "bande": lambda node: node.band_index,
    "mach_bas": lambda node: node.mach_low,
    "mach_haut": lambda node: node.mach_high,
    "braquage": lambda node: node.deflection.name,
    "dl": lambda node: node.deflection.dl,
    "dm": lambda node: node.deflection.dm,
    "dn": lambda node: node.deflection.dn,
    "configuration": lambda node: LIBELLE_CONFIG[node.calc_config],
    "cout_relatif": lambda node: node.relative_cost,
    "composantes_nulles": (
        lambda node: " ".join(node.zero_components) if node.zero_components else "—"
    ),
    "origine": lambda node: str(node.origin),
}


def _valeur(node: DoeNode, cle: str) -> Any:
    """La valeur d'une colonne pour un nœud, déjà traduite en français."""
    calculee = _COLONNES_CALCULEES.get(cle)
    return calculee(node) if calculee is not None else node.values.get(cle)


# --- primitives de mise en forme -------------------------------------------


def _titre(ws: Worksheet, ligne: int, texte: str, sous_titre: str, largeur: int) -> int:
    """Bloc de titre en tête de feuille. Rend la ligne suivante libre."""
    ws.merge_cells(start_row=ligne, start_column=1, end_row=ligne, end_column=largeur)
    cellule = ws.cell(row=ligne, column=1, value=texte)
    cellule.font = TITRE
    cellule.alignment = GAUCHE
    ws.row_dimensions[ligne].height = 24

    ws.merge_cells(start_row=ligne + 1, start_column=1, end_row=ligne + 1, end_column=largeur)
    cellule = ws.cell(row=ligne + 1, column=1, value=sous_titre)
    cellule.font = SOUS_TITRE
    cellule.alignment = GAUCHE
    ws.row_dimensions[ligne + 1].height = 14
    return ligne + 3


def _bloc_cle_valeur(
    ws: Worksheet, ligne: int, titre: str, lignes: Sequence[tuple[str, Any, str]]
) -> int:
    """Un bloc « intitulé / valeur », utilisé par la feuille de synthèse."""
    cellule = ws.cell(row=ligne, column=1, value=titre)
    cellule.font = ENTETE
    cellule.fill = REMPLI_ENTETE
    cellule.alignment = GAUCHE
    cellule.border = BORDURE
    ws.cell(row=ligne, column=2).fill = REMPLI_ENTETE
    ws.cell(row=ligne, column=2).border = BORDURE
    ligne += 1

    for intitule, valeur, format_nombre in lignes:
        gauche = ws.cell(row=ligne, column=1, value=intitule)
        gauche.font = CORPS
        gauche.alignment = GAUCHE
        gauche.border = BORDURE

        droite = ws.cell(row=ligne, column=2, value=valeur)
        droite.font = CORPS_GRAS
        droite.alignment = DROITE if format_nombre != "@" else GAUCHE
        droite.number_format = format_nombre
        droite.border = BORDURE
        ligne += 1

    return ligne + 1


def _mise_en_page(ws: Worksheet, *, paysage: bool, lignes_titre: str = "") -> None:
    """Réglages d'impression : un tableau professionnel s'imprime bien."""
    ws.page_setup.orientation = "landscape" if paysage else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    # « Ajuster à la largeur » n'agit que si la propriété de feuille le déclare ;
    # openpyxl ne crée l'objet que si la feuille en a déjà besoin.
    proprietes = ws.sheet_properties.pageSetUpPr
    if proprietes is None:
        proprietes = PageSetupProperties()
        ws.sheet_properties.pageSetUpPr = proprietes
    proprietes.fitToPage = True

    ws.print_options.horizontalCentered = True
    if lignes_titre:
        ws.print_title_rows = lignes_titre

    pied = ws.oddFooter
    if pied is not None:
        pied.right.text = "Page &P / &N"
        pied.right.size = 8
        pied.left.text = "&F — &A"
        pied.left.size = 8


# --- feuille : synthèse -----------------------------------------------------


def _feuille_synthese(
    wb: Workbook, plan: DoePlan, study: Study, couverture: CoverageResult | None
) -> None:
    """Ce que le chef de projet lit, et rien d'autre."""
    ws = wb.create_sheet("Synthèse")
    ws.column_dimensions["A"].width = 42.0
    ws.column_dimensions["B"].width = 30.0

    ligne = _titre(
        ws,
        1,
        "Plan de calcul CFD",
        f"{study.name} — établi le {datetime.now():%d/%m/%Y}",
        2,
    )

    ligne = _bloc_cle_valeur(
        ws,
        ligne,
        "Volume et coût",
        [
            ("Nombre de cas de calcul", plan.n_nodes, FMT_ENTIER),
            ("Coût total (équivalents configuration complète)", plan.total_cost, FMT_COUT),
            ("Coût sans exploitation des symétries", plan.naive_cost, FMT_COUT),
            ("Économie apportée par les symétries", plan.saving, FMT_PCT),
        ],
    )

    repartition = plan.cost_by_config()
    ligne = _bloc_cle_valeur(
        ws,
        ligne,
        "Répartition par configuration",
        [
            (f"{LIBELLE_CONFIG[config]} — cas", nombre, FMT_ENTIER)
            for config, (nombre, _) in sorted(repartition.items(), key=lambda kv: -kv[1][0])
        ],
    )

    domaine_bas, domaine_haut = plan.envelope.symmetry.fundamental_domain_deg
    fermeture = "]" if plan.envelope.symmetry.domain_is_closed else "["
    ligne = _bloc_cle_valeur(
        ws,
        ligne,
        "Hypothèses",
        [
            ("Groupe de symétrie", str(study.symmetry.group), "@"),
            ("Nature de la configuration", LIBELLE_GROUPE[study.symmetry.group], "@"),
            (
                "Domaine de l'azimut φ",
                f"[{_fr(domaine_bas)}° ; {_fr(domaine_haut)}°{fermeture}",
                "@",
            ),
            (
                "Azimuts calculés",
                " / ".join(f"{_fr(x)}°" for x in azimuth_levels(study.symmetry)),
                "@",
            ),
            ("Méthode de placement des nœuds", str(plan.method), "@"),
            # Une graine est un identifiant, pas une quantité : pas de séparateur.
            ("Graine de reproductibilité", plan.seed, FMT_IDENT),
            ("Longueur de référence [m]", study.reference.length_m, FMT_3),
            ("Bandes de Mach", len(plan.envelope.bands), FMT_ENTIER),
            (
                "Quantiles de l'enveloppe",
                f"{_fr(100 * plan.envelope.spec.q_low, 3)} % / "
                f"{_fr(100 * plan.envelope.spec.q_high, 3)} %",
                "@",
            ),
            ("Marge appliquée aux bornes", plan.envelope.spec.margin, FMT_PCT),
        ],
    )

    if couverture is not None:
        etat = "complète" if couverture.is_complete else "incomplète — voir les points fautifs"
        ligne = _bloc_cle_valeur(
            ws,
            ligne,
            "Contrôle de couverture",
            [
                ("Taux de couverture des trajectoires", couverture.rate, FMT_PCT2),
                ("Points de vol contrôlés", couverture.n_points, FMT_ENTIER),
                ("Points en extrapolation", couverture.n_points - couverture.n_inside, FMT_ENTIER),
                ("État", etat, "@"),
            ],
        )

    note = ws.cell(
        row=ligne,
        column=1,
        value=(
            "Le coût est exprimé en équivalents configuration complète : chaque cas est "
            "compté au prorata du domaine de calcul que sa symétrie autorise."
        ),
    )
    note.font = LEGENDE
    note.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells(start_row=ligne, start_column=1, end_row=ligne + 1, end_column=2)

    _mise_en_page(ws, paysage=False)


# --- feuille : plan de calcul ----------------------------------------------


def _feuille_plan(wb: Workbook, plan: DoePlan, study: Study) -> None:
    """Le tableau de travail : une ligne par cas."""
    ws = wb.create_sheet("Plan de calcul")
    colonnes = _colonnes(plan)
    n = len(colonnes)

    ligne = _titre(
        ws,
        1,
        "Plan de calcul CFD",
        f"{study.name} — {plan.n_nodes} cas — "
        f"{_fr(plan.total_cost)} équivalents configuration complète",
        n,
    )

    # En-têtes groupés : une bande fusionnée par famille de colonnes.
    ligne_groupe = ligne
    debut = 0
    for index in range(n + 1):
        fin_de_groupe = index == n or colonnes[index].groupe != colonnes[debut].groupe
        if fin_de_groupe and index > debut:
            ws.merge_cells(
                start_row=ligne_groupe,
                start_column=debut + 1,
                end_row=ligne_groupe,
                end_column=index,
            )
            cellule = ws.cell(row=ligne_groupe, column=debut + 1, value=colonnes[debut].groupe)
            cellule.font = ENTETE_GROUPE
            cellule.fill = REMPLI_GROUPE
            cellule.alignment = CENTRE
            for numero in range(debut + 1, index + 1):
                ws.cell(row=ligne_groupe, column=numero).border = BORDURE
                ws.cell(row=ligne_groupe, column=numero).fill = REMPLI_GROUPE
            debut = index
    ws.row_dimensions[ligne_groupe].height = 18

    ligne_entete = ligne_groupe + 1
    for index, colonne in enumerate(colonnes, start=1):
        cellule = ws.cell(row=ligne_entete, column=index, value=colonne.entete)
        cellule.font = ENTETE
        cellule.fill = REMPLI_ENTETE
        cellule.alignment = CENTRE
        cellule.border = BORDURE
        ws.column_dimensions[get_column_letter(index)].width = colonne.largeur
    ws.row_dimensions[ligne_entete].height = 30

    premiere_donnee = ligne_entete + 1
    for rang, node in enumerate(plan.nodes):
        rangee = premiere_donnee + rang
        bande_paire = node.band_index % 2 == 0
        for index, colonne in enumerate(colonnes, start=1):
            cellule = ws.cell(row=rangee, column=index, value=_valeur(node, colonne.cle))
            cellule.font = CORPS
            cellule.alignment = colonne.alignement
            cellule.number_format = colonne.format
            cellule.border = BORDURE
            if colonne.cle == "configuration":
                cellule.fill = PatternFill("solid", fgColor=COULEUR_CONFIG[node.calc_config])
            elif not bande_paire:
                cellule.fill = REMPLI_BANDE

    derniere = premiere_donnee + len(plan.nodes) - 1
    if plan.nodes:
        ws.auto_filter.ref = f"A{ligne_entete}:{get_column_letter(n)}{derniere}"
    ws.freeze_panes = ws.cell(row=premiere_donnee, column=5)

    _mise_en_page(ws, paysage=True, lignes_titre=f"{ligne_groupe}:{ligne_entete}")


# --- feuille : enveloppe ----------------------------------------------------


def _feuille_enveloppe(wb: Workbook, plan: DoePlan, study: Study) -> None:
    """La justification du plan : les bornes conditionnelles, bande par bande."""
    ws = wb.create_sheet("Enveloppe")
    entetes = [
        ("Bande", 8.0, FMT_ENTIER, CENTRE),
        ("Mach min", 10.0, FMT_2, DROITE),
        ("Mach max", 10.0, FMT_2, DROITE),
        ("Points", 10.0, FMT_ENTIER, DROITE),
        ("Variable", 16.0, "@", GAUCHE),
        ("Rôle", 20.0, "@", GAUCHE),
        ("Borne basse", 14.0, FMT_4, DROITE),
        ("Quantile bas", 14.0, FMT_4, DROITE),
        ("Médiane", 14.0, FMT_4, DROITE),
        ("Quantile haut", 14.0, FMT_4, DROITE),
        ("Borne haute", 14.0, FMT_4, DROITE),
        ("Niveaux", 9.0, FMT_ENTIER, CENTRE),
    ]

    ligne = _titre(
        ws,
        1,
        "Enveloppe conditionnelle",
        f"{study.name} — bornes recalculées bande de Mach par bande de Mach, "
        f"quantiles {_fr(100 * plan.envelope.spec.q_low, 3)} % / "
        f"{_fr(100 * plan.envelope.spec.q_high, 3)} % élargis d'une marge de "
        f"{_fr(100 * plan.envelope.spec.margin)} %",
        len(entetes),
    )

    ligne_entete = ligne
    for index, (titre, largeur, _, _) in enumerate(entetes, start=1):
        cellule = ws.cell(row=ligne_entete, column=index, value=titre)
        cellule.font = ENTETE
        cellule.fill = REMPLI_ENTETE
        cellule.alignment = CENTRE
        cellule.border = BORDURE
        ws.column_dimensions[get_column_letter(index)].width = largeur
    ws.row_dimensions[ligne_entete].height = 28

    rangee = ligne_entete + 1
    for bande in plan.envelope.bands:
        bande_paire = bande.band.index % 2 == 0
        for variable in bande.variables:
            grandeurs = [
                bande.band.index,
                bande.band.mach_low,
                bande.band.mach_high,
                bande.n_points,
                variable.name,
                LIBELLE_ROLE[variable.spec.role],
                variable.bounds.low,
                variable.bounds.q_low_value,
                variable.bounds.median,
                variable.bounds.q_high_value,
                variable.bounds.high,
                len(variable.levels),
            ]
            # Les cinq bornes partagent un format : la notation scientifique
            # s'impose dès que l'une d'elles est trop longue à écrire en clair.
            bornes = [x for x in grandeurs[6:11] if isinstance(x, float) and x == x]
            grand = max((abs(x) for x in bornes), default=0.0)
            for index, ((_, _, format_nombre, alignement), valeur) in enumerate(
                zip_strict(entetes, grandeurs), start=1
            ):
                cellule = ws.cell(row=rangee, column=index, value=valeur)
                cellule.font = CORPS
                cellule.alignment = alignement
                cellule.number_format = (
                    FMT_SCI
                    if format_nombre == FMT_4 and grand >= SEUIL_SCIENTIFIQUE
                    else format_nombre
                )
                cellule.border = BORDURE
                if not bande_paire:
                    cellule.fill = REMPLI_BANDE
            rangee += 1

    if rangee > ligne_entete + 1:
        ws.auto_filter.ref = f"A{ligne_entete}:{get_column_letter(len(entetes))}{rangee - 1}"
    ws.freeze_panes = ws.cell(row=ligne_entete + 1, column=6)
    _mise_en_page(ws, paysage=True, lignes_titre=f"{ligne_entete}:{ligne_entete}")


# --- feuille : paramètres ---------------------------------------------------


def _feuille_parametres(wb: Workbook, plan: DoePlan, study: Study) -> None:
    """Le rôle de chaque colonne, et la légende des configurations."""
    ws = wb.create_sheet("Paramètres")
    ws.column_dimensions["A"].width = 20.0
    ws.column_dimensions["B"].width = 24.0
    ws.column_dimensions["C"].width = 12.0
    ws.column_dimensions["D"].width = 14.0
    ws.column_dimensions["E"].width = 46.0

    ligne = _titre(
        ws,
        1,
        "Paramètres et légendes",
        f"{study.name} — rôle de chaque colonne, et coût de chaque configuration",
        5,
    )

    for index, titre in enumerate(
        ("Variable", "Rôle", "Niveaux", "Échelle", "Origine du rôle"), start=1
    ):
        cellule = ws.cell(row=ligne, column=index, value=titre)
        cellule.font = ENTETE
        cellule.fill = REMPLI_ENTETE
        cellule.alignment = CENTRE
        cellule.border = BORDURE
    ligne += 1

    for spec in plan.envelope.specs:
        if not spec.is_active:
            continue
        origine = spec.detection if spec.auto else "déclaré dans le fichier d'étude"
        valeurs = [
            spec.display,
            LIBELLE_ROLE[spec.role],
            spec.n_levels,
            str(spec.scale),
            origine,
        ]
        for index, valeur in enumerate(valeurs, start=1):
            cellule = ws.cell(row=ligne, column=index, value=valeur)
            cellule.font = CORPS
            cellule.alignment = GAUCHE if index in (1, 2, 4, 5) else CENTRE
            cellule.number_format = FMT_ENTIER if index == 3 else "@"
            cellule.border = BORDURE
            if spec.auto:
                cellule.fill = REMPLI_ACCENT
        ligne += 1

    ligne += 1
    note = ws.cell(
        row=ligne,
        column=1,
        value=(
            "Les lignes surlignées portent un rôle déduit des valeurs et non déclaré. "
            "Vérifiez-les et figez-les dans la section « parametres » du fichier d'étude."
        ),
    )
    note.font = LEGENDE
    ws.merge_cells(start_row=ligne, start_column=1, end_row=ligne, end_column=5)
    ligne += 2

    for index, titre in enumerate(("Configuration de calcul", "Coût relatif"), start=1):
        cellule = ws.cell(row=ligne, column=index, value=titre)
        cellule.font = ENTETE
        cellule.fill = REMPLI_ENTETE
        cellule.alignment = CENTRE
        cellule.border = BORDURE
    ligne += 1

    from cfd_traj.core.symmetry import RELATIVE_COST

    for config in CalcConfig:
        gauche = ws.cell(row=ligne, column=1, value=LIBELLE_CONFIG[config])
        gauche.font = CORPS
        gauche.alignment = GAUCHE
        gauche.border = BORDURE
        gauche.fill = PatternFill("solid", fgColor=COULEUR_CONFIG[config])

        droite = ws.cell(row=ligne, column=2, value=RELATIVE_COST[config])
        droite.font = CORPS
        droite.alignment = DROITE
        droite.number_format = FMT_3
        droite.border = BORDURE
        ligne += 1

    ligne += 1
    for texte in (
        f"Groupe de symétrie déclaré : {study.symmetry.group} "
        f"({LIBELLE_GROUPE[study.symmetry.group]}).",
        "Sur un plan de miroir, les composantes hors plan CY, Cn et Cl sont nulles par "
        "théorème — non pas petites : nulles.",
        "Un jeu de braquages antisymétrique (roulis, lacet) détruit ce miroir et impose "
        "la configuration complète.",
    ):
        cellule = ws.cell(row=ligne, column=1, value=texte)
        cellule.font = LEGENDE
        cellule.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.merge_cells(start_row=ligne, start_column=1, end_row=ligne, end_column=5)
        ligne += 1

    _mise_en_page(ws, paysage=False)


# --- point d'entrée ---------------------------------------------------------


def write_plan_excel(
    plan: DoePlan,
    study: Study,
    path: str | Path,
    *,
    coverage: CoverageResult | None = None,
) -> Path:
    """Écrit le classeur Excel du plan et rend le chemin produit.

    La symétrie n'est pas un argument : elle est déjà portée par l'enveloppe du
    plan, et la redemander ouvrirait la porte à un classeur incohérent avec le
    plan qu'il décrit.
    """
    cible = Path(path)
    cible.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Workbook() naît avec une feuille vide ; les quatre vraies la remplacent.
    vierge = wb.active
    if vierge is not None:
        wb.remove(vierge)

    _feuille_synthese(wb, plan, study, coverage)
    _feuille_plan(wb, plan, study)
    _feuille_enveloppe(wb, plan, study)
    _feuille_parametres(wb, plan, study)

    wb.properties.title = f"Plan de calcul CFD — {study.name}"
    wb.properties.subject = "Plan d'expériences déduit d'un lot de trajectoires dispersées"
    wb.properties.creator = "cfd-traj"
    wb.properties.description = (
        f"{plan.n_nodes} cas, {_fr(plan.total_cost)} équivalents configuration complète, "
        f"méthode « {plan.method} », graine {plan.seed}."
    )

    wb.active = 0
    wb.save(cible)
    return cible
